"""
Import a PAO-style year-end inventory sheet into legal_cases.

Usage (after pip install -r requirements-migration.txt):
  python scripts/import_inventory_xlsx.py path/to/inventory.xlsx \\
    --token YOUR_JWT

Expects header row with columns like:
  CONTROL NUMBER, PARTY REPRESENTED, TITLE OF THE CASE, COURT/BODY,
  CASE NO., CAUSE OF ACTION, STATUS OF THE CASE, LAST ACTION TAKEN,
  CAUSE OF TERMINATION, DATE OF TERMINATION, address, contact number, ...
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import urllib.error
import urllib.request

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def main():
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", type=Path)
    p.add_argument("--base-url", default="http://127.0.0.1:8000")
    p.add_argument("--token", required=True, help="JWT from POST /api/auth/token")
    args = p.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        print("Install: pip install -r requirements-migration.txt", file=sys.stderr)
        raise SystemExit(1) from e

    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Empty sheet")
        return
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def col(*names):
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    idx_control = col("control number", "control")
    idx_party = col("party represented", "party")
    idx_title = col("title of the", "title")
    idx_court = col("court", "body")
    idx_case_no = col("case no")
    idx_cause = col("cause of action")
    idx_status = col("status of the case", "status")
    idx_last = col("last action")
    idx_term_cause = col("cause of termination")
    idx_term_date = col("date of termination")

    if idx_control is None:
        print("Could not find CONTROL NUMBER column. Header:", header[:20])
        return

    created = 0
    for row in rows[1:]:
        if not row or row[idx_control] is None:
            continue
        control = str(row[idx_control]).strip()
        if not control or control.lower() == "none":
            continue

        payload = {
            "control_number": control,
            "party_represented": _cell(row, idx_party),
            "title_of_case": _cell(row, idx_title),
            "court_body": _cell(row, idx_court),
            "case_number": _cell(row, idx_case_no),
            "cause_of_action": _cell(row, idx_cause),
            "last_action_taken": _cell(row, idx_last),
            "cause_of_termination": _cell(row, idx_term_cause),
            "status": "terminated" if _cell(row, idx_term_date) else "pending",
        }
        dd = _cell(row, idx_term_date)
        if dd:
            payload["date_of_termination"] = _parse_date(dd)

        req = urllib.request.Request(
            f"{args.base_url.rstrip('/')}/api/cases/",
            data=json.dumps(payload).encode(),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {args.token}",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req) as resp:
                if resp.status == 200:
                    created += 1
        except urllib.error.HTTPError as e:
            body = e.read().decode(errors="replace")
            if e.code == 400 and "already exists" in body:
                continue
            print(f"Skip {control}: {e.code} {body[:200]}")

    print(f"Imported (new rows): {created}")


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    return None


if __name__ == "__main__":
    main()
